"""Tests for the ISI-count classifiers that populate REDCap ``*_isi_n`` radios."""

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from parser.mem_parser import (
    ASICF_ISIS,
    A_SICI_ISIS,
    TSICF_ISIS,
    TSICI_ISIS,
    _assign_isi_counts,
    _classify_asicf_isi_n,
    _classify_asici_isi_n,
    _classify_tsicf_isi_n,
    _classify_tsici_isi_n,
    initialize_record,
)
from processing.redcap_mapper import to_redcap_dataframe

import pandas as pd


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _record_with(prefix: str, isi_labels: list[str]) -> dict:
    """Return a fresh record with the given ISI labels populated (value 1.0)."""
    record = initialize_record()
    for label in isi_labels:
        record[f"{prefix}_{label}"] = 1.0
    return record


# ---------------------------------------------------------------------------
# T-SICI
# ---------------------------------------------------------------------------

def test_tsici_none_when_no_data():
    assert _classify_tsici_isi_n(initialize_record()) is None


def test_tsici_three_isi_short_protocol():
    # 1 + 2.5 + 3 ms → option 1
    record = _record_with("T_SICI", ["1.0ms", "2.5ms", "3.0ms"])
    assert _classify_tsici_isi_n(record) == 1


def test_tsici_six_isi_default_protocol():
    record = _record_with(
        "T_SICI", ["1.0ms", "1.5ms", "2ms", "2.5ms", "3.0ms", "3.5ms"]
    )
    assert _classify_tsici_isi_n(record) == 2


def test_tsici_nine_isi_extended_protocol():
    record = _record_with(
        "T_SICI",
        ["1.0ms", "1.5ms", "2ms", "2.5ms", "3.0ms", "3.5ms", "4ms", "5ms", "7ms"],
    )
    assert _classify_tsici_isi_n(record) == 3


# ---------------------------------------------------------------------------
# T-SICF  (option code equals count: 3, 6, 9, 21)
# ---------------------------------------------------------------------------

def test_tsicf_none_when_no_data():
    assert _classify_tsicf_isi_n(initialize_record()) is None


def test_tsicf_fourteen_isi_protocol():
    # 14 ISIs at 1.0..4.9 by 0.3 → option 1 (matches the new template scheme)
    isis_14 = [isi for isi in TSICF_ISIS if float(isi.rstrip("ms")) <= 4.9]
    assert len(isis_14) == 14
    record = _record_with("T_SICF", isis_14)
    assert _classify_tsicf_isi_n(record) == 1


def test_tsicf_nine_isi_low_range():
    # 9 ISIs spanning 1.0..3.4 → option 2
    isis_9 = [isi for isi in TSICF_ISIS if float(isi.rstrip("ms")) <= 3.4]
    assert len(isis_9) == 9
    record = _record_with("T_SICF", isis_9)
    assert _classify_tsicf_isi_n(record) == 2


def test_tsicf_nine_isi_mid_range():
    # 9 ISIs spanning 2.5..4.9 → option 3
    isis_9 = [
        isi for isi in TSICF_ISIS
        if 2.5 <= float(isi.rstrip("ms")) <= 4.9
    ]
    assert len(isis_9) == 9
    record = _record_with("T_SICF", isis_9)
    assert _classify_tsicf_isi_n(record) == 3


# ---------------------------------------------------------------------------
# A-SICI  (dictionary only defines 6-ISI options; default 1 when data present)
# ---------------------------------------------------------------------------

def test_asici_none_when_no_data():
    assert _classify_asici_isi_n(initialize_record()) is None


def test_asici_six_isi_default_protocol():
    # New template gives a_sici_1000_isi_n the same options as t_sici_p_isi_n
    # (1=3 ISIs, 2=6 ISIs, 3=9 ISIs).
    record = _record_with(
        "A_SICI", ["1.0ms", "1.5ms", "2.0ms", "2.5ms", "3.0ms", "3.5ms"]
    )
    assert _classify_asici_isi_n(record) == 2


def test_asici_three_isi_short_protocol():
    record = _record_with("A_SICI", ["1.0ms", "2.5ms", "3.0ms"])
    assert _classify_asici_isi_n(record) == 1


def test_asici_nine_isi_extended_protocol():
    record = _record_with(
        "A_SICI",
        ["1.0ms", "1.5ms", "2.0ms", "2.5ms", "3.0ms", "3.5ms",
         "4.0ms", "5.0ms", "7.0ms"],
    )
    assert _classify_asici_isi_n(record) == 3


# ---------------------------------------------------------------------------
# A-SICF  (range-sensitive)
# ---------------------------------------------------------------------------

def test_asicf_none_when_no_data():
    assert _classify_asicf_isi_n(initialize_record()) is None


def test_asicf_fourteen_isi_1_0_to_4_9():
    # 14 ISIs spanning 1.0..4.9 (the default protocol) → option 1
    isis_14 = [isi for isi in ASICF_ISIS if float(isi.rstrip("ms")) <= 4.9]
    assert len(isis_14) == 14
    record = _record_with("A_SICF", isis_14)
    assert _classify_asicf_isi_n(record) == 1


def test_asicf_nine_isi_low_range():
    # 9 ISIs, max <= 3.4 → option 2
    isis_9 = [isi for isi in ASICF_ISIS if float(isi.rstrip("ms")) <= 3.4]
    assert len(isis_9) == 9
    record = _record_with("A_SICF", isis_9)
    assert _classify_asicf_isi_n(record) == 2


def test_asicf_nine_isi_mid_range():
    # 9 ISIs spanning 2.5..4.9 → option 3
    isis_9 = [
        isi for isi in ASICF_ISIS
        if 2.5 <= float(isi.rstrip("ms")) <= 4.9
    ]
    assert len(isis_9) == 9
    record = _record_with("A_SICF", isis_9)
    assert _classify_asicf_isi_n(record) == 3


# ---------------------------------------------------------------------------
# End-to-end: DataFrame → REDCap
# ---------------------------------------------------------------------------

def test_isi_n_written_as_integer_in_csv(tmp_path):
    """Regression: radio option codes must be written as ``"1"`` not ``"1.0"``.

    REDCap rejects floats for radio fields, so the exporter casts these
    columns to Int64 before writing the CSV.
    """
    import csv as _csv
    from reports.redcap_exporter import generate_redcap_import

    # Build a minimal internal DataFrame with one SNBR record populated
    record = _record_with(
        "T_SICI", ["1.0ms", "1.5ms", "2ms", "2.5ms", "3.0ms", "3.5ms"]
    )
    record["Study"] = "SNBR"
    record["ID"] = 1
    record["Date"] = "15/04/2026"
    record["Stimulated_cortex"] = "L"
    record["source_file"] = "SNBR-001-TP1C50415A.MEM"
    _assign_isi_counts(record)
    py_df = pd.DataFrame([record])

    # Minimal REDCap export CSV: one matching row with empty isi_n
    rc_export = tmp_path / "rc_export.csv"
    with open(rc_export, "w", newline="", encoding="utf-8") as f:
        w = _csv.writer(f)
        w.writerow([
            "record_id", "redcap_event_name", "tt_test_date", "cortex",
            "t_sici_p_isi_n",
        ])
        w.writerow(["1", "visit_1_arm_1", "2026-04-15", "1", ""])

    # Minimal dictionary (only the radio field we care about)
    dict_csv = tmp_path / "dict.csv"
    with open(dict_csv, "w", newline="", encoding="utf-8") as f:
        w = _csv.writer(f)
        w.writerow(["Variable / Field Name", "Form Name"])
        w.writerow(["t_sici_p_isi_n", "tms_values"])

    # Minimal import template (must list any column we emit)
    template_csv = tmp_path / "template.csv"
    with open(template_csv, "w", newline="", encoding="utf-8") as f:
        w = _csv.writer(f)
        w.writerow(["record_id", "redcap_event_name", "t_sici_p_isi_n"])

    import_df, output_path, _ = generate_redcap_import(
        py_df, rc_export, dict_csv, template_csv, tmp_path,
    )

    # Read the written CSV as text to verify the on-disk format
    text = output_path.read_text(encoding="utf-8")
    assert "t_sici_p_isi_n" in text
    # The value row must contain ",2," (integer) and never ",2.0,"
    assert ",2\n" in text or ",2\r\n" in text or text.rstrip().endswith(",2")
    assert ",2.0" not in text


def test_skips_new_participant_ids_not_in_redcap(tmp_path):
    """New participant IDs (absent from the REDCap export) must be dropped entirely."""
    import csv as _csv
    from reports.redcap_exporter import generate_redcap_import

    def _mk_record(pid: int) -> dict:
        r = _record_with(
            "T_SICI", ["1.0ms", "1.5ms", "2ms", "2.5ms", "3.0ms", "3.5ms"]
        )
        r["Study"] = "SNBR"
        r["ID"] = pid
        r["Date"] = "15/04/2026"
        r["Stimulated_cortex"] = "L"
        r["source_file"] = f"SNBR-{pid:03d}-TP1C50415A.MEM"
        _assign_isi_counts(r)
        return r

    # Two participants in the MEM data: ID 1 (known) and ID 999 (new).
    py_df = pd.DataFrame([_mk_record(1), _mk_record(999)])

    rc_export = tmp_path / "rc_export.csv"
    with open(rc_export, "w", newline="", encoding="utf-8") as f:
        w = _csv.writer(f)
        w.writerow([
            "record_id", "redcap_event_name", "tt_test_date", "cortex",
            "t_sici_p_isi_n",
        ])
        # Only ID 1 is registered in REDCap.
        w.writerow(["1", "visit_1_arm_1", "2026-04-15", "1", ""])

    dict_csv = tmp_path / "dict.csv"
    with open(dict_csv, "w", newline="", encoding="utf-8") as f:
        w = _csv.writer(f)
        w.writerow(["Variable / Field Name", "Form Name"])
        w.writerow(["t_sici_p_isi_n", "tms_values"])

    template_csv = tmp_path / "template.csv"
    with open(template_csv, "w", newline="", encoding="utf-8") as f:
        w = _csv.writer(f)
        w.writerow(["record_id", "redcap_event_name", "t_sici_p_isi_n"])

    import_df, _, stats = generate_redcap_import(
        py_df, rc_export, dict_csv, template_csv, tmp_path,
    )

    assert list(import_df["record_id"]) == [1]
    assert 999 not in set(import_df["record_id"])
    assert stats["skipped_new_ids"] == [999]


def test_include_new_ids_toggle_on_adds_new_participant_rows(tmp_path):
    """With ``include_new_ids=True`` the exporter emits rows for new IDs."""
    import csv as _csv
    from reports.redcap_exporter import generate_redcap_import

    def _mk_record(pid: int) -> dict:
        r = _record_with(
            "T_SICI", ["1.0ms", "1.5ms", "2ms", "2.5ms", "3.0ms", "3.5ms"]
        )
        r["Study"] = "SNBR"
        r["ID"] = pid
        r["Date"] = "15/04/2026"
        r["Stimulated_cortex"] = "L"
        r["source_file"] = f"SNBR-{pid:03d}-TP1C50415A.MEM"
        _assign_isi_counts(r)
        return r

    py_df = pd.DataFrame([_mk_record(1), _mk_record(999)])

    rc_export = tmp_path / "rc_export.csv"
    with open(rc_export, "w", newline="", encoding="utf-8") as f:
        w = _csv.writer(f)
        w.writerow([
            "record_id", "redcap_event_name", "tt_test_date", "cortex",
            "t_sici_p_isi_n",
        ])
        w.writerow(["1", "visit_1_arm_1", "2026-04-15", "1", ""])

    dict_csv = tmp_path / "dict.csv"
    with open(dict_csv, "w", newline="", encoding="utf-8") as f:
        w = _csv.writer(f)
        w.writerow(["Variable / Field Name", "Form Name"])
        w.writerow(["t_sici_p_isi_n", "tms_values"])

    template_csv = tmp_path / "template.csv"
    with open(template_csv, "w", newline="", encoding="utf-8") as f:
        w = _csv.writer(f)
        w.writerow(["record_id", "redcap_event_name", "t_sici_p_isi_n"])

    import_df, _, stats = generate_redcap_import(
        py_df, rc_export, dict_csv, template_csv, tmp_path,
        include_new_ids=True,
    )

    ids = set(import_df["record_id"])
    assert 1 in ids and 999 in ids
    assert stats["skipped_new_ids"] == []
    assert stats["new_ids_added"] == [999]
    # The new-participant row should have an empty event name for the user
    # to fill in before import.
    new_row = import_df[import_df["record_id"] == 999].iloc[0]
    assert new_row["redcap_event_name"] == ""


def test_xlsx_change_report_highlights_changed_and_new_cells(tmp_path):
    """The optional xlsx report shows only changed/added cells, with the
    right fills (yellow/green) and rich-text "NEW (OLD)" formatting."""
    import csv as _csv
    from openpyxl import load_workbook
    from openpyxl.cell.rich_text import CellRichText
    from reports.redcap_exporter import generate_redcap_import

    record = _record_with(
        "T_SICI", ["1.0ms", "1.5ms", "2ms", "2.5ms", "3.0ms", "3.5ms"]
    )
    record["Study"] = "SNBR"
    record["ID"] = 1
    record["Date"] = "15/04/2026"
    record["Stimulated_cortex"] = "L"
    record["source_file"] = "SNBR-001-TP1C50415A.MEM"
    # Give T_SICI_1.0ms a distinct numeric value that differs from the REDCap
    # stored value, so the diff marks it as "changed" (yellow).
    record["T_SICI_1.0ms"] = 75.0
    _assign_isi_counts(record)
    py_df = pd.DataFrame([record])

    rc_export = tmp_path / "rc_export.csv"
    with open(rc_export, "w", newline="", encoding="utf-8") as f:
        w = _csv.writer(f)
        w.writerow([
            "record_id", "redcap_event_name", "tt_test_date", "cortex",
            "t_sici_p_1_0_ms", "t_sici_p_isi_n",
        ])
        # Old t_sici_p_1_0_ms = 80 (will change → yellow);
        # t_sici_p_isi_n is blank (will fill → green).
        w.writerow(["1", "visit_1_arm_1", "2026-04-15", "1", "80", ""])

    dict_csv = tmp_path / "dict.csv"
    with open(dict_csv, "w", newline="", encoding="utf-8") as f:
        w = _csv.writer(f)
        w.writerow(["Variable / Field Name", "Form Name"])
        w.writerow(["t_sici_p_1_0_ms", "tms_values"])
        w.writerow(["t_sici_p_isi_n", "tms_values"])

    template_csv = tmp_path / "template.csv"
    with open(template_csv, "w", newline="", encoding="utf-8") as f:
        w = _csv.writer(f)
        w.writerow([
            "record_id", "redcap_event_name",
            "t_sici_p_1_0_ms", "t_sici_p_isi_n",
        ])

    xlsx_dir = tmp_path / "reports"
    _, _, stats = generate_redcap_import(
        py_df, rc_export, dict_csv, template_csv, tmp_path,
        xlsx_report_dir=xlsx_dir,
    )

    assert "xlsx_report_path" in stats
    xlsx_path = Path(stats["xlsx_report_path"])
    assert xlsx_path.exists()

    wb = load_workbook(xlsx_path, rich_text=True)
    ws = wb.active

    # Find the two data columns
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    changed_col = headers.index("t_sici_p_1_0_ms") + 1
    new_col = headers.index("t_sici_p_isi_n") + 1

    changed_cell = ws.cell(row=2, column=changed_col)
    new_cell = ws.cell(row=2, column=new_col)

    # Changed cell: yellow fill + rich text with both values
    assert changed_cell.fill.fgColor.rgb.endswith("FFFF00")
    assert isinstance(changed_cell.value, CellRichText)
    rendered = "".join(tb.text for tb in changed_cell.value)
    assert "75" in rendered and "(80)" in rendered

    # Newly filled cell: green fill, bold plain string (isi_n = 2)
    assert new_cell.fill.fgColor.rgb.endswith("C6EFCE")
    assert str(new_cell.value) == "2"
    assert new_cell.font.bold is True


def test_obsolete_high_end_sici_isis_dropped_from_column_order():
    """The new template removed t_sici_p_*_ms / a_sici_1000_*_ms ISIs above 7ms."""
    from processing.redcap_mapper import REDCAP_COLUMN_ORDER

    for token in ("10_0", "15_0", "20_0", "25_0", "30_0"):
        assert f"t_sici_p_{token}_ms" not in REDCAP_COLUMN_ORDER
        assert f"a_sici_1000_{token}_ms" not in REDCAP_COLUMN_ORDER


def test_tms_coil_extracted_and_mapped_to_radio_code(tmp_path):
    """A 'TMS Coil:\\tMagStim D70^2' header line yields tms_coil = 1 in REDCap."""
    from parser.mem_parser import _extract_tms_coil

    assert _extract_tms_coil("TMS Coil:\tMagStim D70^2") == "D70^2"
    assert _extract_tms_coil("TMS Coil:\tMagStim D70") == "D70"
    assert _extract_tms_coil("TMS Coil:\tDCC") == "DCC"
    assert _extract_tms_coil("TMS Coil:\tunknown brand") is None

    # End-to-end: internal "D70^2" string maps to REDCap option code 1.
    record = initialize_record()
    record["ID"] = 1
    record["TMS_coil"] = "D70^2"
    df = pd.DataFrame([record])
    rc = to_redcap_dataframe(df)
    assert rc["tms_coil"].iloc[0] == 1


def test_completion_gate_emitted_when_protocol_has_value(tmp_path):
    """Auto-set tsicip_completion=1 when a t_sici_p_* value is emitted."""
    import csv as _csv
    from reports.redcap_exporter import generate_redcap_import

    record = _record_with(
        "T_SICI", ["1.0ms", "1.5ms", "2ms", "2.5ms", "3.0ms", "3.5ms"]
    )
    record["Study"] = "SNBR"
    record["ID"] = 1
    record["Date"] = "15/04/2026"
    record["Stimulated_cortex"] = "L"
    record["source_file"] = "SNBR-001-TP1C50415A.MEM"
    _assign_isi_counts(record)
    py_df = pd.DataFrame([record])

    rc_export = tmp_path / "rc_export.csv"
    with open(rc_export, "w", newline="", encoding="utf-8") as f:
        w = _csv.writer(f)
        w.writerow([
            "record_id", "redcap_event_name", "tt_test_date", "cortex",
            "t_sici_p_1_0_ms", "t_sici_p_isi_n", "tsicip_completion",
        ])
        # tsicip_completion blank → exporter should set it to 1.
        w.writerow(["1", "visit_1_arm_1", "2026-04-15", "1", "", "", ""])

    dict_csv = tmp_path / "dict.csv"
    with open(dict_csv, "w", newline="", encoding="utf-8") as f:
        w = _csv.writer(f)
        w.writerow(["Variable / Field Name", "Form Name"])
        for col in ("t_sici_p_1_0_ms", "t_sici_p_isi_n", "tsicip_completion"):
            w.writerow([col, "tms_values"])

    template_csv = tmp_path / "template.csv"
    with open(template_csv, "w", newline="", encoding="utf-8") as f:
        w = _csv.writer(f)
        w.writerow([
            "record_id", "redcap_event_name",
            "t_sici_p_1_0_ms", "t_sici_p_isi_n", "tsicip_completion",
        ])

    import_df, _, _ = generate_redcap_import(
        py_df, rc_export, dict_csv, template_csv, tmp_path,
    )

    assert "tsicip_completion" in import_df.columns
    assert int(import_df["tsicip_completion"].iloc[0]) == 1


def test_completion_gate_skipped_when_already_set_in_redcap(tmp_path):
    """Don't re-emit tsicip_completion=1 if REDCap already has it set."""
    import csv as _csv
    from reports.redcap_exporter import generate_redcap_import

    record = _record_with(
        "T_SICI", ["1.0ms", "1.5ms", "2ms", "2.5ms", "3.0ms", "3.5ms"]
    )
    record["Study"] = "SNBR"
    record["ID"] = 1
    record["Date"] = "15/04/2026"
    record["Stimulated_cortex"] = "L"
    record["source_file"] = "SNBR-001-TP1C50415A.MEM"
    _assign_isi_counts(record)
    record["T_SICI_1.0ms"] = 75.0  # force at least one cell to change
    py_df = pd.DataFrame([record])

    rc_export = tmp_path / "rc_export.csv"
    with open(rc_export, "w", newline="", encoding="utf-8") as f:
        w = _csv.writer(f)
        w.writerow([
            "record_id", "redcap_event_name", "tt_test_date", "cortex",
            "t_sici_p_1_0_ms", "t_sici_p_isi_n", "tsicip_completion",
        ])
        w.writerow(["1", "visit_1_arm_1", "2026-04-15", "1", "80", "2", "1"])

    dict_csv = tmp_path / "dict.csv"
    with open(dict_csv, "w", newline="", encoding="utf-8") as f:
        w = _csv.writer(f)
        w.writerow(["Variable / Field Name", "Form Name"])
        for col in ("t_sici_p_1_0_ms", "t_sici_p_isi_n", "tsicip_completion"):
            w.writerow([col, "tms_values"])

    template_csv = tmp_path / "template.csv"
    with open(template_csv, "w", newline="", encoding="utf-8") as f:
        w = _csv.writer(f)
        w.writerow([
            "record_id", "redcap_event_name",
            "t_sici_p_1_0_ms", "t_sici_p_isi_n", "tsicip_completion",
        ])

    import_df, _, _ = generate_redcap_import(
        py_df, rc_export, dict_csv, template_csv, tmp_path,
    )

    # tsicip_completion shouldn't be in the row at all (already 1 in REDCap).
    if "tsicip_completion" in import_df.columns:
        assert pd.isna(import_df["tsicip_completion"].iloc[0])


def test_redcap_dataframe_emits_isi_n_columns():
    record = _record_with(
        "T_SICI", ["1.0ms", "1.5ms", "2ms", "2.5ms", "3.0ms", "3.5ms"]
    )
    record["ID"] = 1
    _assign_isi_counts(record)

    df = pd.DataFrame([record])
    rc = to_redcap_dataframe(df)

    assert "t_sici_p_isi_n" in rc.columns
    assert "t_sicf_p_isi_n" in rc.columns
    assert "a_sici_1000_isi_n" in rc.columns
    assert "a_sicf_isi_n" in rc.columns
    assert rc["t_sici_p_isi_n"].iloc[0] == 2
