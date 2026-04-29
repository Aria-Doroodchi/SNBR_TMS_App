"""
Generate REDCap import CSVs by comparing Python-parsed TMS data against
the current REDCap export.

This module is the backend engine used by both the GUI workflow and the
standalone ``scripts/generate_redcap_import.py`` script.

Public API
----------
find_latest_dated_file(directory, prefix) -> Path
generate_redcap_import(py_dataframe, redcap_data_csv,
                       redcap_dict_csv, redcap_template_csv,
                       output_dir) -> (DataFrame, Path, dict)
"""

from __future__ import annotations

import csv
import re
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd

from processing.redcap_mapper import (
    REDCAP_COLUMN_ORDER,
    REDCAP_RADIO_INT_COLS,
    to_redcap_dataframe,
)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

TOLERANCE = 0.01

REQUIRED_KEY_COLS = ["record_id", "redcap_event_name"]

TMS_VALUE_COLS = list(REDCAP_COLUMN_ORDER)

# REDCap cortex encoding: 1 = Left, 2 = Right
_RC_CORTEX_TO_PY = {"1": "L", "2": "R", "1.0": "L", "2.0": "R"}

# Regex to find YYYY-MM-DD in a filename
_DATE_PATTERN = re.compile(r"(\d{4}-\d{2}-\d{2})")

OUTPUT_STEM = "SNBR_TMS_import"
XLSX_REPORT_STEM = "SNBR_TMS_changes"

# Highlight colours for the xlsx change report
_FILL_YELLOW = "FFFF00"  # changed cells
_FILL_GREEN = "C6EFCE"   # newly added cells


# ---------------------------------------------------------------------------
# File discovery
# ---------------------------------------------------------------------------

def find_latest_dated_file(directory: str | Path, prefix: str) -> Path:
    """Find the most recent date-stamped CSV matching *prefix* in *directory*.

    Looks for files named ``{prefix}YYYY-MM-DD*.csv`` and returns the one
    with the most recent date.  Raises ``FileNotFoundError`` with a clear
    message if no matching files exist.
    """
    dir_path = Path(directory)
    if not dir_path.is_dir():
        raise FileNotFoundError(
            f"Directory does not exist: {dir_path}"
        )

    candidates: list[tuple[datetime, Path]] = []
    for csv_file in dir_path.glob(f"{prefix}*.csv"):
        match = _DATE_PATTERN.search(csv_file.name)
        if match:
            try:
                file_date = datetime.strptime(match.group(1), "%Y-%m-%d")
                candidates.append((file_date, csv_file))
            except ValueError:
                continue

    if not candidates:
        raise FileNotFoundError(
            f"No date-stamped CSV files matching '{prefix}*.csv' "
            f"found in {dir_path}"
        )

    candidates.sort(key=lambda t: t[0], reverse=True)
    return candidates[0][1]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _parse_py_date(d: str) -> str | None:
    try:
        return datetime.strptime(str(d).strip(), "%d/%m/%Y").strftime("%Y-%m-%d")
    except (ValueError, TypeError):
        return None


def _normalise_cortex_to_py(val) -> str | None:
    s = str(val).strip()
    if s in _RC_CORTEX_TO_PY:
        return _RC_CORTEX_TO_PY[s]
    if s.upper() in ("L", "R"):
        return s.upper()
    return None


def _is_valid_number(val) -> bool:
    try:
        v = float(val)
        return np.isfinite(v)
    except (ValueError, TypeError):
        return False


def _validate_against_template(
    import_df: pd.DataFrame,
    template_path: Path,
) -> list[str]:
    """Check that all import columns exist in the REDCap import template."""
    with open(template_path, "r", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        template_cols = set(next(reader))

    bad = [c for c in import_df.columns if c not in template_cols]
    return bad


def _validate_against_dictionary(
    import_df: pd.DataFrame,
    dict_path: Path,
) -> list[str]:
    """Check that TMS value columns are in the data dictionary as tms_values."""
    warnings: list[str] = []
    dd_fields: set[str] = set()
    with open(dict_path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            if row.get("Form Name", "") == "tms_values":
                dd_fields.add(row["Variable / Field Name"])

    tms_in_import = [
        c for c in import_df.columns if c not in REQUIRED_KEY_COLS
    ]
    for col in tms_in_import:
        if col not in dd_fields:
            warnings.append(f"Column '{col}' not in tms_values data dictionary")

    return warnings


# ---------------------------------------------------------------------------
# Change-report xlsx (human-readable diff, parallel to the import CSV)
# ---------------------------------------------------------------------------

def _format_xlsx_value(val) -> str:
    """Render a numeric import value as a short string for the xlsx report."""
    if val is None or (isinstance(val, float) and not np.isfinite(val)):
        return ""
    if isinstance(val, (int, np.integer)):
        return str(int(val))
    try:
        f = float(val)
    except (TypeError, ValueError):
        return str(val)
    if f == int(f):
        return str(int(f))
    return f"{f:.2f}".rstrip("0").rstrip(".")


def _write_change_report_xlsx(
    import_rows: list[dict],
    old_value_rows: list[dict],
    output_dir: Path,
    timestamp: str,
) -> Path:
    """Write an xlsx report showing only changed/added cells.

    Formatting:
      * Changed cells (old value present): yellow fill, "NEW (OLD)" where
        NEW is bold and OLD is regular.
      * New cells (old value missing): green fill, value in bold.
      * Only rows and columns that contain at least one change are kept.

    Returns the written file path.
    """
    from openpyxl import Workbook
    from openpyxl.cell.rich_text import CellRichText, TextBlock
    from openpyxl.cell.text import InlineFont
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    yellow = PatternFill("solid", fgColor=_FILL_YELLOW)
    green = PatternFill("solid", fgColor=_FILL_GREEN)
    bold = Font(bold=True)
    bold_inline = InlineFont(b=True)
    plain_inline = InlineFont()

    # Figure out which TMS columns have at least one change
    tms_cols_present: list[str] = []
    seen: set[str] = set()
    for row in import_rows:
        for col in row:
            if col in REQUIRED_KEY_COLS or col in seen:
                continue
            seen.add(col)
            tms_cols_present.append(col)
    tms_cols_present = [c for c in TMS_VALUE_COLS if c in seen]

    header = REQUIRED_KEY_COLS + tms_cols_present

    wb = Workbook()
    ws = wb.active
    ws.title = "Changes"

    # Header row
    for col_idx, name in enumerate(header, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = bold
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for r_idx, (new_row, old_row) in enumerate(
        zip(import_rows, old_value_rows), start=2
    ):
        for c_idx, col in enumerate(header, start=1):
            ws_cell = ws.cell(row=r_idx, column=c_idx)
            if col in REQUIRED_KEY_COLS:
                ws_cell.value = new_row.get(col, "")
                continue

            new_val = new_row.get(col)
            if new_val is None or (
                isinstance(new_val, float) and not np.isfinite(new_val)
            ):
                continue  # unchanged cell — leave blank, no fill

            new_str = _format_xlsx_value(new_val)
            old_val = old_row.get(col)
            is_newly_filled = old_val is None

            if is_newly_filled:
                ws_cell.value = new_str
                ws_cell.font = bold
                ws_cell.fill = green
            else:
                old_str = _format_xlsx_value(old_val)
                ws_cell.value = CellRichText(
                    TextBlock(bold_inline, new_str),
                    TextBlock(plain_inline, f" ({old_str})"),
                )
                ws_cell.fill = yellow

    # Auto-size columns roughly
    for c_idx, col in enumerate(header, start=1):
        max_len = max(
            [len(str(col))]
            + [
                len(str(ws.cell(row=r, column=c_idx).value or ""))
                for r in range(2, ws.max_row + 1)
            ]
        )
        ws.column_dimensions[get_column_letter(c_idx)].width = min(
            max(10, max_len + 2), 40
        )

    ws.freeze_panes = "C2"

    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"{XLSX_REPORT_STEM}_{timestamp}.xlsx"
    wb.save(output_path)
    return output_path


# ---------------------------------------------------------------------------
# Main engine
# ---------------------------------------------------------------------------

def generate_redcap_import(
    py_dataframe: pd.DataFrame,
    redcap_data_csv: Path,
    redcap_dict_csv: Path,
    redcap_template_csv: Path,
    output_dir: Path,
    *,
    include_new_ids: bool = False,
    xlsx_report_dir: Path | str | None = None,
) -> tuple[pd.DataFrame, Path, dict]:
    """Generate a REDCap import CSV with only changed TMS values.

    Parameters
    ----------
    py_dataframe : pd.DataFrame
        Internal DataFrame from ``df_builder`` (full schema, all studies).
    redcap_data_csv : Path
        Current REDCap data export CSV.
    redcap_dict_csv : Path
        REDCap data dictionary CSV.
    redcap_template_csv : Path
        REDCap import template CSV.
    output_dir : Path
        Directory where the timestamped import CSV will be written.

    Returns
    -------
    (import_df, output_path, summary)
        import_df : pd.DataFrame — the import data
        output_path : Path — where the CSV was written
        summary : dict — statistics and quality check results
    """
    # -- 1. Prepare Python data ------------------------------------------------
    py_snbr = py_dataframe[py_dataframe["Study"] == "SNBR"].copy()
    py_snbr["ID"] = pd.to_numeric(py_snbr["ID"], errors="coerce").astype("Int64")
    py_snbr["date_iso"] = py_snbr["Date"].apply(_parse_py_date)
    py_snbr["cortex_py"] = py_snbr["Stimulated_cortex"].apply(
        lambda v: str(v).strip().upper() if pd.notna(v) else None
    )

    # Deduplicate for REDCap: keep only the first MEM file per (ID, date).
    # The suffix character (last char of filename stem, e.g. "A", "B", "C")
    # determines file ordering within the same participant visit.
    py_snbr["_file_suffix"] = py_snbr["source_file"].apply(
        lambda f: Path(str(f)).stem[-1].upper() if pd.notna(f) else "Z"
    )
    py_snbr = (
        py_snbr.sort_values("_file_suffix", ascending=True)
        .drop_duplicates(subset=["ID", "date_iso"], keep="first")
        .drop(columns=["_file_suffix"])
    )

    py_rc = to_redcap_dataframe(py_snbr)
    py_rc["_record_id"] = pd.to_numeric(
        py_rc["record_id"], errors="coerce"
    ).astype("Int64")
    py_rc["_date_iso"] = py_snbr["date_iso"].values
    py_rc["_cortex_py"] = py_snbr["cortex_py"].values

    # -- 2. Load REDCap data ---------------------------------------------------
    rc_df = pd.read_csv(redcap_data_csv, low_memory=False)
    rc_df["record_id"] = pd.to_numeric(
        rc_df["record_id"], errors="coerce"
    ).astype("Int64")
    rc_df["tt_test_date"] = rc_df["tt_test_date"].astype(str).str.strip()
    rc_df["_cortex_py"] = (
        rc_df["cortex"].astype(str).str.strip().apply(_normalise_cortex_to_py)
    )

    shared_tms_cols = [
        c for c in TMS_VALUE_COLS if c in py_rc.columns and c in rc_df.columns
    ]

    # Participant-ID filter: by default, only emit rows for IDs already
    # registered in REDCap. When ``include_new_ids`` is True the filter is
    # skipped and new participants flow through as additional import rows.
    known_ids = set(rc_df["record_id"].dropna().unique())
    py_ids_all = set(py_rc["_record_id"].dropna().unique())
    new_ids_detected = sorted(
        int(pid) for pid in (py_ids_all - known_ids) if pd.notna(pid)
    )
    if include_new_ids:
        skipped_new_ids: list[int] = []
    else:
        skipped_new_ids = new_ids_detected
        py_rc = py_rc[py_rc["_record_id"].isin(known_ids)].copy()

    # -- 3. Match and diff -----------------------------------------------------
    import_rows: list[dict] = []
    # Parallel list tracking the OLD REDCap value for each changed cell
    # (None for newly-filled cells). Used by the xlsx change report.
    old_value_rows: list[dict] = []
    stats = {
        "matched": 0,
        "rows_changed": 0,
        "cells_changed": 0,
        "cells_filled": 0,
        "per_column": {},
        "per_participant": [],
        "skipped_new_ids": skipped_new_ids,
        "new_ids_added": [],
    }

    for _, py_row in py_rc.iterrows():
        pid = py_row["_record_id"]
        py_date = py_row.get("_date_iso", "")
        py_cortex = py_row.get("_cortex_py", "")
        if pd.isna(pid) or not py_date or not py_cortex:
            continue

        mask = (
            (rc_df["record_id"] == pid)
            & (rc_df["tt_test_date"] == py_date)
            & (rc_df["_cortex_py"] == py_cortex)
        )
        rc_match = rc_df[mask]
        is_new_participant = int(pid) in new_ids_detected

        if len(rc_match) == 0 and not is_new_participant:
            continue

        if len(rc_match) > 0:
            stats["matched"] += 1
            rc_row = rc_match.iloc[0]
            event_name = rc_row["redcap_event_name"]
        else:
            # New participant — no existing REDCap row. Emit the parsed values
            # with an empty event_name so the user can fill it before import.
            rc_row = None
            event_name = ""
            if int(pid) not in stats["new_ids_added"]:
                stats["new_ids_added"].append(int(pid))

        changed: dict[str, float | int | str] = {}
        old_vals: dict[str, float | int | None] = {}
        for col in shared_tms_cols:
            py_val = py_row.get(col)
            rc_val = rc_row.get(col) if rc_row is not None else None
            py_has = _is_valid_number(py_val)
            rc_has = _is_valid_number(rc_val)
            is_radio_int = col in REDCAP_RADIO_INT_COLS

            def _format(v: float) -> float | int:
                return int(round(float(v))) if is_radio_int else round(float(v), 2)

            if py_has and rc_has:
                if abs(float(py_val) - float(rc_val)) > TOLERANCE:
                    changed[col] = _format(py_val)
                    old_vals[col] = _format(rc_val)
                    stats["per_column"][col] = (
                        stats["per_column"].get(col, 0) + 1
                    )
                    stats["cells_changed"] += 1
            elif py_has and not rc_has:
                changed[col] = _format(py_val)
                old_vals[col] = None
                stats["per_column"][col] = (
                    stats["per_column"].get(col, 0) + 1
                )
                stats["cells_filled"] += 1

        if changed:
            row = {"record_id": int(pid), "redcap_event_name": event_name}
            row.update(changed)
            import_rows.append(row)
            old_value_rows.append({
                "record_id": int(pid),
                "redcap_event_name": event_name,
                **old_vals,
            })
            stats["rows_changed"] += 1

    # -- 4. Build DataFrame & deduplicate --------------------------------------
    if not import_rows:
        empty = pd.DataFrame(columns=REQUIRED_KEY_COLS)
        warn_list: list[str] = []
        if skipped_new_ids:
            warn_list.append(
                f"Skipped {len(skipped_new_ids)} new participant ID(s) "
                f"not yet in REDCap: {skipped_new_ids}"
            )
        stats["quality_checks"] = {
            "status": "PASS" if not warn_list else "WARNINGS",
            "warnings": warn_list,
        }
        return empty, Path(), stats

    import_df = pd.DataFrame(import_rows)

    # Merge duplicate (record_id, event) rows
    key = ["record_id", "redcap_event_name"]
    if import_df.duplicated(subset=key, keep=False).any():
        merged_rows = []
        for (pid, evt), group in import_df.groupby(key, sort=False):
            merged = {"record_id": pid, "redcap_event_name": evt}
            for col in import_df.columns:
                if col in key:
                    continue
                vals = group[col].dropna()
                if len(vals) > 0:
                    merged[col] = vals.iloc[0]
            merged_rows.append(merged)
        import_df = pd.DataFrame(merged_rows)

    # Final column ordering
    tms_present = [c for c in TMS_VALUE_COLS if c in import_df.columns]
    import_df = import_df[REQUIRED_KEY_COLS + tms_present]

    # Cast radio/integer columns to nullable Int64 so the CSV writes "1", not
    # "1.0" (REDCap rejects floats for radio fields).
    for col in REDCAP_RADIO_INT_COLS:
        if col in import_df.columns:
            import_df[col] = pd.to_numeric(
                import_df[col], errors="coerce"
            ).astype("Int64")

    # Per-participant summary
    for _, row in import_df.iterrows():
        n = sum(1 for c in tms_present if pd.notna(row.get(c)))
        stats["per_participant"].append({
            "record_id": int(row["record_id"]),
            "event": row["redcap_event_name"],
            "columns_changed": n,
        })

    # -- 5. Quality checks -----------------------------------------------------
    qc_warnings: list[str] = []

    # Validate against template
    template_bad = _validate_against_template(import_df, redcap_template_csv)
    if template_bad:
        qc_warnings.append(
            f"Columns not in import template: {', '.join(template_bad)}"
        )

    # Validate against dictionary
    dict_warns = _validate_against_dictionary(import_df, redcap_dict_csv)
    qc_warnings.extend(dict_warns)

    # Surface new-participant handling (informational, not a failure).
    if skipped_new_ids:
        qc_warnings.append(
            f"Skipped {len(skipped_new_ids)} new participant ID(s) "
            f"not yet in REDCap: {skipped_new_ids}"
        )
    if stats["new_ids_added"]:
        qc_warnings.append(
            f"Added {len(stats['new_ids_added'])} new participant ID(s) "
            f"with EMPTY redcap_event_name — fill in the event name before "
            f"import: {stats['new_ids_added']}"
        )

    # Check all event names are valid
    valid_events = set(rc_df["redcap_event_name"].dropna().unique())
    bad_events = [
        e for e in import_df["redcap_event_name"] if e not in valid_events
    ]
    if bad_events:
        qc_warnings.append(f"Unknown event names: {bad_events}")

    # Check no non-finite values
    all_vals = import_df[tms_present].values.flatten()
    non_nan = all_vals[~pd.isna(all_vals)]
    non_finite = sum(1 for v in non_nan if not np.isfinite(float(v)))
    if non_finite:
        qc_warnings.append(f"{non_finite} non-finite values detected")

    stats["quality_checks"] = {
        "status": "PASS" if not qc_warnings else "WARNINGS",
        "warnings": qc_warnings,
    }

    # -- 6. Write output -------------------------------------------------------
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    output_path = output_dir / f"{OUTPUT_STEM}_{timestamp}.csv"
    import_df.to_csv(output_path, index=False)

    # Optional: write a human-readable xlsx change report
    if xlsx_report_dir is not None and import_rows:
        xlsx_path = _write_change_report_xlsx(
            import_rows=import_rows,
            old_value_rows=old_value_rows,
            output_dir=Path(xlsx_report_dir),
            timestamp=timestamp,
        )
        stats["xlsx_report_path"] = str(xlsx_path)

    return import_df, output_path, stats
